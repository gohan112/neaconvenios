"""
NeaEvento — app web para el evento de empresa.

Dos caras:
  · Participante: abre su ENLACE PERSONAL (/p/<código>) y ve su equipo, sus
    compañeros, el programa del día y los lugares; confirma su asistencia.
    Al abrirlo por primera vez queda registrado («activado»).
  · Organización: panel en /admin (contraseña opcional con la variable
    EVENTO_ADMIN_PASSWORD) para participantes, equipos + sorteo, agenda,
    lugares, enlaces y datos del evento.

Ejecutar:  python app.py   →  http://localhost:8502
Variables: EVENTO_ADMIN_PASSWORD (contraseña del panel; OBLIGATORIA en internet),
           EVENTO_DB_PATH (ruta del evento.db), PUERTO (por defecto 8502).
"""

from __future__ import annotations

import csv
import io
import os
import re
from functools import wraps

from urllib.parse import quote

from flask import (Flask, Response, abort, flash, get_flashed_messages, jsonify,
                   redirect, request, send_file, session)

import db
import paginas

CARPETA = os.path.dirname(os.path.abspath(__file__))
PUERTO = int(os.environ.get("PUERTO", "8502"))

app = Flask(__name__, static_folder=os.path.join(CARPETA, "assets"),
            static_url_path="/assets")

db.iniciar()
app.secret_key = db.secreto_app()
app.config.update(SESSION_COOKIE_HTTPONLY=True, SESSION_COOKIE_SAMESITE="Lax")

PASSWORD_ADMIN = (os.environ.get("EVENTO_ADMIN_PASSWORD") or "").strip()


# ------------------------------------------------------------------ utilidades

def _avisos():
    return get_flashed_messages(with_categories=True)


def _sin_password() -> bool:
    return not PASSWORD_ADMIN


def requiere_admin(funcion):
    @wraps(funcion)
    def envoltura(*args, **kwargs):
        if PASSWORD_ADMIN and not session.get("admin"):
            return redirect("/admin/entrar")
        return funcion(*args, **kwargs)
    return envoltura


def _entero_o_none(valor) -> int | None:
    try:
        return int(valor)
    except (TypeError, ValueError):
        return None


def _hora_si_es_hoy(cfg: dict) -> str | None:
    """La hora actual «HH:MM» solo si el evento es HOY (para marcar «AHORA»)."""
    if (cfg.get("fecha") or "").strip() != db.hoy().isoformat():
        return None
    return db.ahora().split(" ")[-1]


def _url_base() -> tuple[str, bool]:
    """(url_base, definida). Si no está configurada, usa la del navegador."""
    configurada = (db.leer_config().get("url_base") or "").strip().rstrip("/")
    if configurada:
        return configurada, True
    return request.url_root.rstrip("/"), False


def _enlace_de(token: str, url_base: str) -> str:
    return f"{url_base}/p/{token}"


def _telefono_wa(telefono: str) -> str:
    """Deja solo dígitos; a los móviles españoles de 9 cifras les antepone 34."""
    digitos = re.sub(r"\D", "", telefono or "")
    if len(digitos) == 9 and digitos[0] in "67":
        digitos = "34" + digitos
    return digitos


def _mensaje_whatsapp(cfg: dict, nombre: str, enlace: str, apodo: str = "") -> str:
    plantilla = cfg.get("msg_whatsapp") or ""
    apodo = apodo or (nombre.split() or [""])[0]
    try:
        return plantilla.format(nombre=nombre, enlace=enlace, apodo=apodo)
    except (KeyError, IndexError, ValueError):
        return f"¡Hola, {nombre}! Este es tu enlace personal para el evento: {enlace}"


# ================================================================== público

@app.get("/")
def portada():
    return paginas.render_portada(db.leer_config(), db.hoy())


@app.post("/ir")
def ir_con_codigo():
    codigo = (request.form.get("codigo") or "").strip().split("/")[-1]
    return redirect(f"/p/{codigo}" if codigo else "/")


@app.get("/p/<token>")
def ver_participante(token: str):
    p = db.participante_por_token(token)
    if not p:
        return paginas.render_no_encontrado(db.leer_config()), 404
    db.marcar_visto(p["id"])

    # Con equipo asignado y sorteo aún no visto → animación del sorteo (una vez)
    if p["equipo_id"] and not p.get("revelado_en"):
        equipos = db.listar_equipos()
        indice = next((i for i, eq in enumerate(equipos)
                       if eq["id"] == p["equipo_id"]), None)
        if indice is not None:
            cfg = db.leer_config()
            datos = db.resumen()
            historia = paginas.lineas_historia(
                cfg, n_equipos=datos["equipos"],
                n_participantes=datos["participantes"])
            return paginas.render_sorteo(cfg=cfg, p=p, equipos=equipos,
                                         indice_final=indice, historia=historia,
                                         avisos=_avisos())

    equipo = db.equipo(p["equipo_id"]) if p["equipo_id"] else None
    companeros = db.miembros(p["equipo_id"]) if p["equipo_id"] else []
    cfg = db.leer_config()
    id_lugar_escape = _entero_o_none(cfg.get("escape_lugar_id"))
    lugar_escape = db.lugar(id_lugar_escape) if id_lugar_escape else None
    id_lugar_karts = _entero_o_none(cfg.get("karts_lugar_id"))
    lugar_karts = db.lugar(id_lugar_karts) if id_lugar_karts else None
    clasif = db.clasificacion()
    final = db.estado_final()
    # Al cerrar la Olimpiada, cada uno ve si tiene premio nada más entrar
    premios = db.ganadores(clasif) if (cfg.get("resultado_final") or "").strip() else None
    return paginas.render_participante(
        cfg=cfg, p=p, equipo=equipo, companeros=companeros,
        agenda=db.agenda_para(p["equipo_id"]), lugares=db.listar_lugares(),
        referencia=db.hoy(), lugar_escape=lugar_escape,
        lugar_karts=lugar_karts, clasif=clasif,
        hora_actual=_hora_si_es_hoy(cfg), corre_final=p["id"] in final["ids"],
        pasa_por_tiempo=any(x["id"] == p["id"] for x in final["por_tiempo"]),
        faltan_tiempos=len(final["pendientes"]),
        ganadores=premios, avisos=_avisos(),
    )


@app.get("/p/<token>/equipo.json")
def equipo_json(token: str):
    """Estado del equipo del participante, para refrescar «Mi equipo» en directo."""
    p = db.participante_por_token(token)
    if not p or not p["equipo_id"]:
        abort(404)
    miembros = db.miembros(p["equipo_id"])
    equipo = db.equipo(p["equipo_id"]) or {}
    color = equipo.get("color") or "#CC0C18"
    dentro = [{"n": paginas.nombre_corto(m), "ini": paginas.iniciales(m),
               "yo": m["id"] == p["id"],
               "cap": m["id"] == equipo.get("capitan_id")}
              for m in miembros if m.get("revelado_en")]
    pendientes = len(miembros) - len(dentro)
    return jsonify(dentro=dentro, pendientes=pendientes, color=color,
                   tinta=paginas.color_texto(color),
                   final=p["id"] in db.estado_final()["ids"],
                   texto=paginas._texto_contador(len(dentro), pendientes))


@app.post("/p/<token>/tiempo_escape")
def capitan_tiempo_escape(token: str):
    """El capitán del equipo apunta la hora de salida de su sala."""
    p = db.participante_por_token(token)
    if not p or not p["equipo_id"]:
        abort(404)
    equipo = db.equipo(p["equipo_id"])
    if not equipo or equipo.get("capitan_id") != p["id"]:
        abort(403)
    texto = (request.form.get("tiempo") or "").strip()
    if texto and db.parsear_hora_dia(texto) is None:
        flash("La hora no se entiende: usa el formato 10:05 (o 10:05:30).", "error")
    else:
        db.poner_tiempo_escape(equipo["id"], texto)
        flash("¡Hora de salida guardada! Ya cuenta en la clasificación.", "ok")
    return redirect(f"/p/{token}#puntos")  # vuelve a la pestaña de puntos


@app.get("/p/<token>/puntos.html")
def puntos_fragmento(token: str):
    """La clasificación en HTML, para refrescarla en directo en la pestaña Puntos."""
    p = db.participante_por_token(token)
    if not p:
        abort(404)
    return paginas.fragmento_clasificacion(db.clasificacion(), nota=False)


@app.post("/p/<token>/tiempo_karts")
def participante_tiempo_karts(token: str):
    """Cada piloto apunta su propia vuelta (la de su tanda o la de la final)."""
    p = db.participante_por_token(token)
    if not p:
        abort(404)
    if "tiempo_final" in request.form and not db.corre_la_final(p):
        abort(403)          # el hueco de la final solo lo tiene quien la corre
    guardados, con_texto, error = 0, 0, False
    for campo, es_final in (("tiempo", False), ("tiempo_final", True)):
        if campo not in request.form:
            continue
        texto = (request.form.get(campo) or "").strip()
        if texto and db.parsear_tiempo_vuelta(texto) is None:
            error = True
            continue        # los que sí se entienden se guardan igual
        db.poner_tiempo_karts(p["id"], texto, final=es_final)
        guardados += 1
        con_texto += 1 if texto else 0
    if error:
        flash("Ese tiempo no se entiende: escríbelo como 1:02.451 (minutos:segundos"
              ".milésimas) o 48.123 si bajaste del minuto.", "error")
    elif con_texto:
        flash("¡Tiempo guardado! Ya cuenta en la clasificación.", "ok")
    elif guardados:
        flash("Tiempo borrado.", "ok")
    return redirect(f"/p/{token}#puntos")


@app.post("/p/<token>/revelado")
def marcar_revelado(token: str):
    """La animación del sorteo ha terminado (o se saltó sin JavaScript)."""
    p = db.participante_por_token(token)
    if not p:
        abort(404)
    db.marcar_revelado(p["id"])
    return redirect(f"/p/{token}")


@app.post("/p/<token>/asistencia")
def confirmar_asistencia(token: str):
    p = db.participante_por_token(token)
    if not p:
        abort(404)
    valor = request.form.get("valor")
    if valor not in ("si", "no"):
        abort(400)
    db.poner_asistencia(p["id"], viene=(valor == "si"))
    flash("¡Gracias! Hemos guardado tu respuesta." if valor == "si"
          else "Respuesta guardada. ¡Una pena que no puedas venir!", "ok")
    return redirect(f"/p/{token}")


@app.get("/salud")
def salud():
    return jsonify(ok=True)


@app.after_request
def sin_guardar_en_cache(respuesta):
    """El móvil tiene que ver siempre lo de ahora, no una copia guardada.

    Sin esto, el navegador (o el visor de enlaces de WhatsApp) puede enseñar
    una página vieja: dices «¡Sí, voy!» y sigues viendo el botón, o el equipo
    y la clasificación se quedan a medias. Los ficheros de /assets no entran:
    Flask ya les pone sus propias cabeceras."""
    respuesta.headers.setdefault("Cache-Control", "no-store, max-age=0")
    return respuesta


# ================================================================== admin: sesión

@app.route("/admin/entrar", methods=["GET", "POST"])
def admin_entrar():
    if _sin_password() or session.get("admin"):
        return redirect("/admin")
    if request.method == "POST":
        import secrets as _secrets
        if _secrets.compare_digest(request.form.get("password", ""), PASSWORD_ADMIN):
            session["admin"] = True
            return redirect("/admin")
        flash("Contraseña incorrecta.", "error")
    return paginas.render_entrar(avisos=_avisos())


@app.get("/admin/salir")
def admin_salir():
    session.pop("admin", None)
    return redirect("/")


# ================================================================== admin: resumen

@app.get("/admin")
@requiere_admin
def admin_resumen():
    return paginas.render_resumen(
        cfg=db.leer_config(), datos=db.resumen(), equipos=db.listar_equipos(),
        referencia=db.hoy(), avisos=_avisos(), sin_password=_sin_password(),
    )


# ================================================================== admin: participantes

@app.get("/admin/participantes")
@requiere_admin
def admin_participantes():
    return paginas.render_participantes(
        lista=db.listar_participantes(), equipos=db.listar_equipos(),
        avisos=_avisos(), sin_password=_sin_password(),
    )


@app.post("/admin/participantes/nuevo")
@requiere_admin
def admin_participante_nuevo():
    nombre = (request.form.get("nombre") or "").strip()
    if not nombre:
        flash("El nombre es obligatorio.", "error")
    else:
        db.crear_participante(
            nombre=nombre,
            apodo=request.form.get("apodo", ""),
            rol=request.form.get("rol", ""),
            telefono=request.form.get("telefono", ""),
            email=request.form.get("email", ""),
            equipo_id=_entero_o_none(request.form.get("equipo_id")),
        )
        flash(f"Añadido «{nombre}». Su enlace personal está en la pestaña Enlaces.", "ok")
    return redirect("/admin/participantes")


def _filas_de_texto_csv(texto: str) -> list[dict]:
    muestra = texto[:3000]
    separador = max([";", ",", "\t"], key=muestra.count)
    lector = csv.reader(io.StringIO(texto), delimiter=separador)
    filas_crudas = [[(celda or "").strip() for celda in fila] for fila in lector]
    filas_crudas = [f for f in filas_crudas if any(f)]
    if not filas_crudas:
        return []

    indices = {"nombre": 0, "telefono": 1, "email": 2, "equipo": 3}
    primera = [c.lower() for c in filas_crudas[0]]
    tiene_cabecera = any("nombre" in c for c in primera)
    if tiene_cabecera:
        indices = {}
        for i, celda in enumerate(primera):
            if "apodo" in celda and "apodo" not in indices:
                indices["apodo"] = i
            elif "nombre" in celda and "nombre" not in indices:
                indices["nombre"] = i
            elif ("tel" in celda or "móvil" in celda or "movil" in celda) and "telefono" not in indices:
                indices["telefono"] = i
            elif ("mail" in celda or "correo" in celda) and "email" not in indices:
                indices["email"] = i
            elif "equipo" in celda and "equipo" not in indices:
                indices["equipo"] = i
            elif (("rol" in celda or "perfil" in celda or "comercial" in celda
                   or "tecnico" in celda or "técnico" in celda)
                  and "rol" not in indices):
                indices["rol"] = i
        filas_crudas = filas_crudas[1:]

    def valor(fila: list[str], campo: str) -> str:
        i = indices.get(campo)
        return fila[i] if i is not None and i < len(fila) else ""

    return [
        {campo: valor(fila, campo)
         for campo in ("nombre", "apodo", "rol", "telefono", "email", "equipo")}
        for fila in filas_crudas
    ]


def _filas_de_fichero(nombre_fichero: str, datos: bytes) -> list[dict]:
    if nombre_fichero.lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        libro = load_workbook(io.BytesIO(datos), read_only=True, data_only=True)
        hoja = libro.worksheets[0]
        lineas = []
        for fila in hoja.iter_rows(values_only=True):
            celdas = ["" if c is None else str(c).strip() for c in fila]
            lineas.append(";".join(celdas))
        libro.close()
        return _filas_de_texto_csv("\n".join(lineas))
    try:
        texto = datos.decode("utf-8-sig")
    except UnicodeDecodeError:
        texto = datos.decode("latin-1")
    return _filas_de_texto_csv(texto)


@app.post("/admin/participantes/importar")
@requiere_admin
def admin_participantes_importar():
    fichero = request.files.get("fichero")
    if not fichero or not fichero.filename:
        flash("Elige un fichero .xlsx o .csv.", "error")
        return redirect("/admin/participantes")
    try:
        filas = _filas_de_fichero(fichero.filename, fichero.read())
    except Exception as exc:  # noqa: BLE001 — mostrar el motivo, no tragarlo
        flash(f"No se pudo leer el fichero: {exc}", "error")
        return redirect("/admin/participantes")
    nuevos, existentes = db.importar(filas)
    aviso = f"Importados {nuevos} participante(s) nuevo(s)."
    if existentes:
        aviso += (f" Otros {existentes} ya existían: se conservan tal cual "
                  f"(solo se les rellenan los datos que tuvieran vacíos).")
    if not nuevos and not existentes:
        aviso = ("No se encontró ningún nombre en el fichero. Comprueba que la primera "
                 "columna (o una columna «nombre») tiene los nombres.")
    flash(aviso, "ok" if (nuevos or existentes) else "error")
    return redirect("/admin/participantes")


@app.post("/admin/participantes/pegar")
@requiere_admin
def admin_participantes_pegar():
    lineas = (request.form.get("lineas") or "").splitlines()
    filas = []
    for linea in lineas:
        partes = [p.strip() for p in re.split(r"[;\t]", linea)]
        if partes and partes[0]:
            filas.append({
                "nombre": partes[0],
                "telefono": partes[1] if len(partes) > 1 else "",
                "email": partes[2] if len(partes) > 2 else "",
                "equipo": partes[3] if len(partes) > 3 else "",
                "apodo": "",
                "rol": "",
            })
    nuevos, existentes = db.importar(filas)
    aviso = f"Añadidos {nuevos} participante(s)."
    if existentes:
        aviso += f" Otros {existentes} ya existían y se conservan tal cual."
    flash(aviso, "ok")
    return redirect("/admin/participantes")


@app.get("/admin/participantes/<int:participante_id>")
@requiere_admin
def admin_participante_editar(participante_id: int):
    p = db.participante(participante_id)
    if not p:
        abort(404)
    url_base, _definida = _url_base()
    return paginas.render_participante_editar(
        p=p, equipos=db.listar_equipos(), enlace=_enlace_de(p["token"], url_base),
        avisos=_avisos(), sin_password=_sin_password(),
    )


@app.post("/admin/participantes/<int:participante_id>/guardar")
@requiere_admin
def admin_participante_guardar(participante_id: int):
    if not db.participante(participante_id):
        abort(404)
    nombre = (request.form.get("nombre") or "").strip()
    if not nombre:
        flash("El nombre es obligatorio.", "error")
        return redirect(f"/admin/participantes/{participante_id}")
    db.editar_participante(
        participante_id, nombre=nombre,
        apodo=request.form.get("apodo", ""),
        rol=request.form.get("rol", ""),
        telefono=request.form.get("telefono", ""),
        email=request.form.get("email", ""),
        equipo_id=_entero_o_none(request.form.get("equipo_id")),
        notas=request.form.get("notas", ""),
        tanda=request.form.get("tanda", ""),
    )
    flash("Guardado.", "ok")
    return redirect("/admin/participantes")


@app.post("/admin/participantes/<int:participante_id>/borrar")
@requiere_admin
def admin_participante_borrar(participante_id: int):
    db.borrar_participante(participante_id)
    flash("Participante borrado.", "ok")
    return redirect("/admin/participantes")


@app.post("/admin/participantes/<int:participante_id>/token")
@requiere_admin
def admin_participante_token(participante_id: int):
    if not db.participante(participante_id):
        abort(404)
    db.regenerar_token(participante_id)
    flash("Enlace nuevo generado: el anterior ya no funciona. Reenvíaselo.", "ok")
    return redirect(f"/admin/participantes/{participante_id}")


# ================================================================== admin: equipos

@app.get("/admin/equipos")
@requiere_admin
def admin_equipos():
    equipos = db.listar_equipos()
    miembros = {eq["id"]: db.miembros(eq["id"]) for eq in equipos}
    return paginas.render_equipos(
        equipos=equipos, miembros_por_equipo=miembros,
        n_sin_equipo=db.resumen()["sin_equipo"],
        grupos=db.grupos_juntos(), participantes=db.listar_participantes(),
        avisos=_avisos(), sin_password=_sin_password(),
    )


@app.post("/admin/juntos/nuevo")
@requiere_admin
def admin_juntos_nuevo():
    ids = [i for i in (_entero_o_none(x) for x in request.form.getlist("ids"))
           if i is not None]
    if len(ids) < 2:
        flash("Marca al menos a 2 personas para unirlas.", "error")
    else:
        db.crear_grupo_juntos(ids)
        flash(f"Regla creada: esas {len(ids)} personas irán al mismo equipo.", "ok")
    return redirect("/admin/equipos")


@app.post("/admin/juntos/<grupo>/borrar")
@requiere_admin
def admin_juntos_borrar(grupo: str):
    db.deshacer_grupo_juntos(grupo)
    flash("Regla deshecha: ya no tienen por qué ir juntos.", "ok")
    return redirect("/admin/equipos")


@app.post("/admin/equipos/nuevo")
@requiere_admin
def admin_equipos_nuevo():
    nombres = [n for n in (request.form.get("nombres") or "").split(",") if n.strip()]
    creados = db.crear_equipos(nombres)
    flash(f"Creado(s) {creados} equipo(s)." if creados
          else "No se creó ningún equipo (¿nombres vacíos o repetidos?).",
          "ok" if creados else "error")
    return redirect("/admin/equipos")


@app.post("/admin/equipos/<int:equipo_id>/guardar")
@requiere_admin
def admin_equipo_guardar(equipo_id: int):
    if not db.equipo(equipo_id):
        abort(404)
    db.editar_equipo(
        equipo_id, nombre=request.form.get("nombre", ""),
        color=request.form.get("color", ""), emoji=request.form.get("emoji", ""),
        descripcion=request.form.get("descripcion", ""),
    )
    flash("Equipo guardado.", "ok")
    return redirect("/admin/equipos")


@app.post("/admin/equipos/<int:equipo_id>/borrar")
@requiere_admin
def admin_equipo_borrar(equipo_id: int):
    db.borrar_equipo(equipo_id)
    flash("Equipo borrado. Sus miembros han quedado sin equipo.", "ok")
    return redirect("/admin/equipos")


@app.post("/admin/capitanes/sortear")
@requiere_admin
def admin_capitanes_sortear():
    n = db.sortear_capitanes()
    flash(f"Capitanes sorteados: {n}. Cada capitán podrá meter la hora de salida "
          f"de su sala desde su móvil.", "ok")
    return redirect("/admin/equipos")


@app.post("/admin/sorteo")
@requiere_admin
def admin_sorteo():
    try:
        repartidos = db.sortear(todos=request.form.get("modo") == "todos")
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect("/admin/equipos")
    flash(f"Sorteo hecho: {repartidos} participante(s) repartido(s)." if repartidos
          else "No había nadie pendiente de repartir.", "ok")
    return redirect("/admin/equipos")


# ================================================================== admin: agenda

@app.get("/admin/agenda")
@requiere_admin
def admin_agenda():
    cfg = db.leer_config()
    return paginas.render_agenda(
        items=db.listar_agenda(), lugares=db.listar_lugares(),
        equipos=db.listar_equipos(), cfg=cfg,
        participantes=db.listar_participantes(),
        salas=db.parsear_salas(cfg.get("escape_salas", "")),
        avisos=_avisos(), sin_password=_sin_password(),
    )


@app.post("/admin/tandas/sortear")
@requiere_admin
def admin_tandas_sortear():
    n1, n2, n3 = db.sortear_tandas()
    flash(f"Tandas sorteadas: {n1} en la 1ª, {n2} en la 2ª y {n3} en la 3ª "
          f"(a la 3ª se sumarán los mejores tiempos).", "ok")
    return redirect("/admin/agenda")


@app.post("/admin/tandas/deshacer")
@requiere_admin
def admin_tandas_deshacer():
    db.deshacer_tandas()
    flash("Tandas deshechas.", "ok")
    return redirect("/admin/agenda")


@app.post("/admin/salas/config")
@requiere_admin
def admin_salas_config():
    db.guardar_config({
        "escape_titulo": request.form.get("escape_titulo", "Escape room"),
        "escape_hora": db.normalizar_hora(request.form.get("escape_hora", "")),
        "escape_salas": request.form.get("escape_salas", ""),
        "escape_lugar_id": request.form.get("escape_lugar_id", ""),
    })
    flash("Datos de la escape room guardados.", "ok")
    return redirect("/admin/agenda")


@app.post("/admin/salas/sortear")
@requiere_admin
def admin_salas_sortear():
    salas = db.parsear_salas(db.leer_config().get("escape_salas", ""))
    try:
        db.sortear_salas(
            salas,
            sala_excluida=request.form.get("sala_excluida", ""),
            equipo_excluido=_entero_o_none(request.form.get("equipo_excluido")),
        )
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect("/admin/agenda")
    reparto = ", ".join(f"{eq['nombre']} → {eq['sala']}" for eq in db.listar_equipos())
    flash(f"Salas sorteadas: {reparto}.", "ok")
    return redirect("/admin/agenda")


@app.post("/admin/salas/deshacer")
@requiere_admin
def admin_salas_deshacer():
    db.deshacer_salas()
    flash("Reparto de salas deshecho.", "ok")
    return redirect("/admin/agenda")


@app.post("/admin/tandas/config")
@requiere_admin
def admin_tandas_config():
    db.guardar_config({
        "karts_nombre": request.form.get("karts_nombre", "Karts"),
        "karts_hora1": db.normalizar_hora(request.form.get("karts_hora1", "")),
        "karts_hora2": db.normalizar_hora(request.form.get("karts_hora2", "")),
        "karts_hora3": db.normalizar_hora(request.form.get("karts_hora3", "")),
        "karts_lugar_id": request.form.get("karts_lugar_id", ""),
    })
    flash("Horas de las tandas guardadas.", "ok")
    return redirect("/admin/agenda")


def _datos_actividad() -> dict:
    return {
        "hora": request.form.get("hora", ""),
        "hora_fin": request.form.get("hora_fin", ""),
        "actividad": (request.form.get("actividad") or "").strip(),
        "descripcion": request.form.get("descripcion", ""),
        "lugar_id": _entero_o_none(request.form.get("lugar_id")),
        "equipo_id": _entero_o_none(request.form.get("equipo_id")),
    }


@app.post("/admin/agenda/nueva")
@requiere_admin
def admin_agenda_nueva():
    datos = _datos_actividad()
    if not datos["hora"] or not datos["actividad"]:
        flash("Hace falta al menos la hora y el nombre de la actividad.", "error")
    else:
        db.crear_actividad(**datos)
        flash("Actividad añadida.", "ok")
    return redirect("/admin/agenda")


@app.get("/admin/agenda/<int:actividad_id>")
@requiere_admin
def admin_agenda_editar(actividad_id: int):
    a = db.actividad(actividad_id)
    if not a:
        abort(404)
    return paginas.render_agenda_editar(
        a=a, lugares=db.listar_lugares(), equipos=db.listar_equipos(),
        avisos=_avisos(), sin_password=_sin_password(),
    )


@app.post("/admin/agenda/<int:actividad_id>/guardar")
@requiere_admin
def admin_agenda_guardar(actividad_id: int):
    if not db.actividad(actividad_id):
        abort(404)
    datos = _datos_actividad()
    if not datos["hora"] or not datos["actividad"]:
        flash("Hace falta al menos la hora y el nombre de la actividad.", "error")
        return redirect(f"/admin/agenda/{actividad_id}")
    db.editar_actividad(actividad_id, **datos)
    flash("Actividad guardada.", "ok")
    return redirect("/admin/agenda")


@app.post("/admin/agenda/<int:actividad_id>/borrar")
@requiere_admin
def admin_agenda_borrar(actividad_id: int):
    db.borrar_actividad(actividad_id)
    flash("Actividad borrada.", "ok")
    return redirect("/admin/agenda")


# ================================================================== admin: lugares

@app.get("/admin/lugares")
@requiere_admin
def admin_lugares():
    return paginas.render_lugares(lugares=db.listar_lugares(), avisos=_avisos(),
                                  sin_password=_sin_password())


@app.post("/admin/lugares/nuevo")
@requiere_admin
def admin_lugar_nuevo():
    nombre = (request.form.get("nombre") or "").strip()
    if not nombre:
        flash("El nombre del lugar es obligatorio.", "error")
    else:
        db.crear_lugar(nombre=nombre, direccion=request.form.get("direccion", ""),
                       maps=request.form.get("maps", ""),
                       notas=request.form.get("notas", ""))
        flash("Lugar añadido.", "ok")
    return redirect("/admin/lugares")


@app.get("/admin/lugares/<int:lugar_id>")
@requiere_admin
def admin_lugar_editar(lugar_id: int):
    lugar_ = db.lugar(lugar_id)
    if not lugar_:
        abort(404)
    return paginas.render_lugar_editar(lugar_=lugar_, avisos=_avisos(),
                                       sin_password=_sin_password())


@app.post("/admin/lugares/<int:lugar_id>/guardar")
@requiere_admin
def admin_lugar_guardar(lugar_id: int):
    if not db.lugar(lugar_id):
        abort(404)
    nombre = (request.form.get("nombre") or "").strip()
    if not nombre:
        flash("El nombre del lugar es obligatorio.", "error")
        return redirect(f"/admin/lugares/{lugar_id}")
    db.editar_lugar(lugar_id, nombre=nombre,
                    direccion=request.form.get("direccion", ""),
                    maps=request.form.get("maps", ""),
                    notas=request.form.get("notas", ""))
    flash("Lugar guardado.", "ok")
    return redirect("/admin/lugares")


@app.post("/admin/lugares/<int:lugar_id>/borrar")
@requiere_admin
def admin_lugar_borrar(lugar_id: int):
    db.borrar_lugar(lugar_id)
    flash("Lugar borrado.", "ok")
    return redirect("/admin/lugares")


# ================================================================== admin: puntos

@app.get("/admin/puntos")
@requiere_admin
def admin_puntos():
    clasif = db.clasificacion()
    return paginas.render_puntos(
        cfg=db.leer_config(), clasif=clasif, ganadores=db.ganadores(clasif),
        final=db.estado_final(),
        equipos=db.listar_equipos(), participantes=db.listar_participantes(),
        avisos=_avisos(), sin_password=_sin_password(),
    )


@app.post("/admin/puntos/escape")
@requiere_admin
def admin_puntos_escape():
    for eq in db.listar_equipos():
        campo = f"tiempo_{eq['id']}"
        if campo in request.form:
            db.poner_tiempo_escape(eq["id"], request.form.get(campo, ""))
    db.guardar_config({"puntos_escape": request.form.get("puntos_escape", "")})
    flash("Tiempos de la escape room guardados.", "ok")
    return redirect("/admin/puntos")


@app.post("/admin/puntos/karts")
@requiere_admin
def admin_puntos_karts():
    finalistas = set(request.form.getlist("finalista"))
    for p in db.listar_participantes():
        campo = f"tiempo_{p['id']}"
        if campo in request.form:
            db.poner_tiempo_karts(p["id"], request.form.get(campo, ""))
        campo_final = f"final_{p['id']}"
        if campo_final in request.form:
            db.poner_tiempo_karts(p["id"], request.form.get(campo_final, ""), final=True)
        db.marcar_finalista(p["id"], str(p["id"]) in finalistas)
    flash("Tiempos de karts guardados.", "ok")
    return redirect("/admin/puntos")


@app.post("/admin/puntos/final")
@requiere_admin
def admin_puntos_final():
    """Cierra (o reabre) la Olimpiada: la felicitación a los ganadores."""
    cerrar = request.form.get("valor") == "1"
    db.guardar_config({"resultado_final": "1" if cerrar else ""})
    if cerrar:
        premios = db.ganadores(db.clasificacion())
        nombres = ", ".join(x["nombre"] for x in premios["equipos"]) or "nadie todavía"
        flash(f"¡Resultado publicado! Gana {nombres}. Cada uno lo ve al abrir su "
              f"enlace, con el aviso de recoger el premio en los postres.", "ok")
    else:
        flash("Resultado retirado: se deja de felicitar a nadie.", "ok")
    return redirect("/admin/puntos")


# ================================================================== admin: enlaces

def _filas_enlaces(url_base: str) -> list[dict]:
    cfg = db.leer_config()
    filas = []
    for p in db.listar_participantes():
        enlace = _enlace_de(p["token"], url_base)
        telefono_wa = _telefono_wa(p["telefono"])
        wa = ""
        if telefono_wa:
            mensaje = _mensaje_whatsapp(cfg, p["nombre"], enlace,
                                        apodo=p.get("apodo") or "")
            wa = f"https://wa.me/{telefono_wa}?text={quote(mensaje)}"
        filas.append({"nombre": p["nombre"], "apodo": p.get("apodo") or "",
                      "rol": p.get("rol") or "", "telefono": p["telefono"],
                      "email": p["email"], "equipo": p.get("equipo_nombre") or "",
                      "tanda": p.get("tanda") or "",
                      "confirmado": p["confirmado"], "visto_en": p.get("visto_en"),
                      "enlace": enlace, "wa": wa})
    return filas


@app.get("/admin/enlaces")
@requiere_admin
def admin_enlaces():
    url_base, definida = _url_base()
    filas = _filas_enlaces(url_base)
    texto_todos = "\n".join(f"{f['nombre']}: {f['enlace']}" for f in filas)
    return paginas.render_enlaces(
        filas_datos=filas, url_base=url_base, url_definida=definida,
        texto_todos=texto_todos, avisos=_avisos(), sin_password=_sin_password(),
    )


@app.get("/admin/enlaces.csv")
@requiere_admin
def admin_enlaces_csv():
    url_base, _definida = _url_base()
    salida = io.StringIO()
    # ';' y BOM para que el Excel en español lo abra en columnas directamente
    escritor = csv.writer(salida, delimiter=";")
    escritor.writerow(["Nombre", "Apodo", "Rol", "Teléfono", "Email", "Equipo",
                       "Tanda karts", "Enlace personal", "Asistencia",
                       "Abrió el enlace"])
    estados = {1: "Viene", -1: "No viene", 0: "Sin responder"}
    for f in _filas_enlaces(url_base):
        escritor.writerow([f["nombre"], f["apodo"], f["rol"], f["telefono"],
                           f["email"], f["equipo"], f["tanda"],
                           f["enlace"], estados.get(f["confirmado"], ""),
                           f["visto_en"] or ""])
    datos = salida.getvalue().encode("utf-8-sig")
    return Response(datos, mimetype="text/csv; charset=utf-8", headers={
        "Content-Disposition": "attachment; filename=enlaces_evento.csv"})


# ================================================================== admin: evento

@app.get("/admin/evento")
@requiere_admin
def admin_evento():
    return paginas.render_evento(cfg=db.leer_config(), avisos=_avisos(),
                                 sin_password=_sin_password())


@app.post("/admin/evento/guardar")
@requiere_admin
def admin_evento_guardar():
    db.guardar_config({
        "nombre": request.form.get("nombre", ""),
        "fecha": request.form.get("fecha", ""),
        "hora": request.form.get("hora", ""),
        "descripcion": request.form.get("descripcion", ""),
        "historia": request.form.get("historia", ""),
        "contacto": request.form.get("contacto", ""),
        "url_base": (request.form.get("url_base", "") or "").rstrip("/"),
        "msg_whatsapp": request.form.get("msg_whatsapp", ""),
    })
    flash("Datos del evento guardados.", "ok")
    return redirect("/admin/evento")


# ================================================================== copia de seguridad

@app.get("/admin/copia.db")
@requiere_admin
def admin_copia():
    """Descarga el evento entero (un único fichero SQLite)."""
    con = db.conexion()
    con.execute("PRAGMA wal_checkpoint(TRUNCATE)")  # volcar el WAL al fichero
    con.close()
    nombre = f"evento_copia_{db.hoy().strftime('%Y%m%d')}.db"
    return send_file(db.RUTA_DB, as_attachment=True, download_name=nombre,
                     mimetype="application/octet-stream")


@app.post("/admin/restaurar")
@requiere_admin
def admin_restaurar():
    """Sustituye el evento por una copia .db subida (validándola antes)."""
    fichero = request.files.get("fichero")
    if not fichero or not fichero.filename:
        flash("Elige el fichero .db de la copia.", "error")
        return redirect("/admin/evento")
    datos = fichero.read()
    if not datos.startswith(b"SQLite format 3\x00"):
        flash("Ese fichero no es una copia de NeaEvento (.db).", "error")
        return redirect("/admin/evento")
    # Comprobar que abre y que tiene las tablas de la app
    import sqlite3 as _sqlite3
    import tempfile as _tempfile
    with _tempfile.NamedTemporaryFile(delete=False, suffix=".db") as tmp:
        tmp.write(datos)
        ruta_tmp = tmp.name
    try:
        con = _sqlite3.connect(ruta_tmp)
        tablas = {f[0] for f in con.execute(
            "SELECT name FROM sqlite_master WHERE type='table'")}
        con.close()
        if not {"config", "participantes", "equipos"} <= tablas:
            flash("El fichero no contiene los datos de un evento.", "error")
            return redirect("/admin/evento")
    except _sqlite3.Error:
        flash("No se pudo leer el fichero: no parece una copia válida.", "error")
        return redirect("/admin/evento")
    finally:
        os.unlink(ruta_tmp)
    # Sustituir la base de datos (y limpiar los ficheros auxiliares del WAL)
    for sufijo in ("-wal", "-shm"):
        try:
            os.unlink(db.RUTA_DB + sufijo)
        except FileNotFoundError:
            pass
    with open(db.RUTA_DB, "wb") as destino:
        destino.write(datos)
    db.iniciar()  # aplicar migraciones/claves que falten a la copia
    datos_resumen = db.resumen()
    flash(f"Copia restaurada: {datos_resumen['participantes']} participante(s), "
          f"{datos_resumen['equipos']} equipo(s), "
          f"{datos_resumen['actividades']} actividad(es).", "ok")
    return redirect("/admin")


# ================================================================== arranque

if __name__ == "__main__":
    print(f"NeaEvento en http://localhost:{PUERTO}  (panel: /admin)")
    if _sin_password():
        print("AVISO: panel de administración SIN contraseña "
              "(define EVENTO_ADMIN_PASSWORD antes de publicarla en internet).")
    try:
        from waitress import serve
        serve(app, host="0.0.0.0", port=PUERTO, threads=8)
    except ImportError:  # sin waitress (instalación mínima): servidor de desarrollo
        app.run(host="0.0.0.0", port=PUERTO)
